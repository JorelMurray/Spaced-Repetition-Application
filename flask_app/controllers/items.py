from flask_app import app
from flask import Flask, render_template, request, redirect, session, flash
from flask_app.models.user import User
from flask_app.models.project import Project
from flask_app.models.item import Item
from openpyxl import Workbook
import getpass
from flask_bcrypt import Bcrypt
import webbrowser
bcrypt = Bcrypt(app)

confidenceRef = {
    0 : "New",
    1 : "Very Low",
    2 : "Low",
    3 : "Medium",
    4 : "High",
    5 : "Very High"
}
difficultyRef = {
    1 : "Very Easy",
    2 : "Easy",
    3 : "Medium",
    4 : "Hard",
    5 : "Very Hard"
}

@app.route('/additem/<int:pID>')
def addItem(pID):
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')     
    

    return render_template("newitem.html", pID = pID)

@app.route('/createitem', methods = ['POST'])
def createItem():
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')

    isValid = Item.validate(request.form)

    if not isValid:
        return redirect(f"/additem/{request.form['pID']}")

    data = {
        "itemName" : request.form['itemName'],
        "category" : request.form['category'],
        "difficultyLevel" : request.form['difficultyLevel'],
        "itemURL" : request.form['itemURL'],
        "pID" : request.form['pID']
    }

    Item.addItem(data)

    return redirect(f"/viewproject/{request.form['pID']}/")

@app.route('/importitems', methods = ['POST'])
def importItems():
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')

    f = request.files['file']

    pID = request.form['pID']

    if f.filename != '':
        pID = Item.importItems(request.form['pID'], f, difficultyRef )


    return redirect (f"/viewproject/{pID}/")
    
@app.route('/edititem/<int:pID>/<int:itemID>/')
def editItem(pID, itemID):
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')

    data = {
        "id" : itemID
    }
    

    return render_template("edititem.html", currentItem = Item.getOne(data), pID = pID)

@app.route('/updateitem', methods = ['POST'])
def updateItem():
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')
    
    isValid = Item.validate(request.form)

    if not isValid:
        return redirect(f"/edititem/{request.form['pID']}/{request.form['itemID']}")

    data = {
        "id" : request.form['itemID'],
        "itemName" : request.form['itemName'],
        "category" : request.form['category'],
        "difficultyLevel" : request.form['difficultyLevel'],
        "itemURL" : request.form['itemURL']
    }

    Item.updateItem(data)

    return redirect(f"/viewproject/{request.form['pID']}/")

@app.route('/reviewitem/<int:pID>/<int:itemID>')
def reviewItem(pID,itemID):
    if 'userId' not in session:
        flash('Please log in')
        return redirect('/')

    data = {
        "id" : itemID
    }

    return render_template("reviewitem.html", currentItem = Item.getOne(data), pID = pID, confidenceRef = confidenceRef, difficultyRef = difficultyRef )

@app.route('/submitreview', methods = ['POST'])
def submitReview():
    if request.form['confidenceLevel'] == "0":
        flash("Please select a confidence level")
        return redirect (f"/reviewitem/{request.form['pID']}/{request.form['itemID']}")

    data = {
        "id" : request.form['itemID'],
        "confidenceLevel" : request.form['confidenceLevel'],
    }
    Item.reviewItem(data)
    Item.bookAttempt(data)

    return redirect (f"/viewproject/{request.form['pID']}/")

@app.route('/attemptitem/<int:pID>/<int:itemID>')
def attemptItem(pID,itemID):
    data = {
        'id': itemID,
        'pID' : pID
    }
    
    numOpenItems = Item.getOpenItemCount(data)

    if numOpenItems >= 3:
        flash("You can only have 3 items in progress at a time.")
        return redirect(f"/viewproject/{pID}/") 

    Item.attemptItem(data) 
    Item.bookAttempt(data) 

    currentItem = Item.getOne(data)

    if currentItem.itemURL != "":
        webbrowser.open_new_tab(f'{currentItem.itemURL}')
    
    return redirect(f"/viewproject/{pID}/")

@app.route('/deleteitem/<int:pID>/<int:itemID>')
def deleteItem(pID, itemID):
    data = {
        'id': itemID,
    }
    Item.deleteItem(data)

    return redirect(f"/viewproject/{pID}/")

@app.route('/loadtemplate/<int:pID>')
def generateLoadTemplat(pID):
    wb = Workbook()

    ws = wb.active

    ws.title = "Item Load Template"

    ws['A1'] = "Item"
    ws['B1'] = "Category"
    ws['C1'] = "Difficulty"
    ws['D1'] = "URL"

    wb.save(f'C:/Users/{getpass.getuser()}/Downloads/LoadTemplate.xlsx')
    
    return redirect(f"/viewproject/{pID}/")
